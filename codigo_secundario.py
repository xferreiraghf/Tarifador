import os
import patoolib
import pandas as pd
import PyPDF2
from PyPDF2 import PdfReader
import re
import csv
import openpyxl
import shutil

def extrair_arquivo_rar(arquivo_rar, pasta_destino):
    patoolib.extract_archive(arquivo_rar, outdir=pasta_destino)
    print(f"Arquivo {arquivo_rar} descompactado com sucesso em {pasta_destino}")
    for root, dirs, files in os.walk(pasta_destino):
        for file in files:
            if file.endswith(".txt"):
                caminho_arquivo_txt = os.path.join(root, file)
                return caminho_arquivo_txt

def converter_para_xlsx(caminho_arquivo_txt):
    dados = pd.read_csv(caminho_arquivo_txt, sep=';')
    dados_agrupados = dados.groupby('NumNF', as_index=False).agg({'Valor': 'sum', 'MesRef': 'first'})
    colunas_desejadas = ['NumNF', 'MesRef', 'Valor']
    dados_agrupados = dados_agrupados[colunas_desejadas]
    caminho_arquivo_xlsx = os.path.splitext(caminho_arquivo_txt)[0] + '.xlsx'
    dados_agrupados.to_excel(caminho_arquivo_xlsx, index=False)
    novo_nome = os.path.join(os.path.dirname(caminho_arquivo_xlsx), "fatura.xlsx")
    os.rename(caminho_arquivo_xlsx, novo_nome)
    print(f'O arquivo Excel foi criado em: {novo_nome}')
    return novo_nome

def extract_cnpj(text):
    pattern = re.compile(r'(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})')
    match = re.findall(pattern, text)
    return match

def extract_faturas_cnpj_from_pdf(pdf_path, faturas):
    with open(pdf_path, 'rb') as pdf_file:
        reader = PdfReader(pdf_file)
        text = ''
        for page in reader.pages:
            text += page.extract_text()
    faturas_encontradas = []
    for fatura in faturas:
        match = re.search(r"NFST {} CNPJ: (\d{{2}}\.\d{{3}}\.\d{{3}}/\d{{4}}-\d{{2}})".format(re.escape(fatura)), text)
        if match:
            faturas_encontradas.append((fatura, match.group(1)))
    return faturas_encontradas

def extrair_valores(pdf_path):
    with open(pdf_path, 'rb') as pdf_file:
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        valores_totais = {}
        
        for page_num in range(len(pdf_reader.pages)):
            text = pdf_reader.pages[page_num].extract_text()

            match_total = re.search(r'TOTAL TIM S\.A.*?(\d{1,3}(?:\.\d{3})*,\d+)', text)
            if match_total:
                valor_total = match_total.group(1).replace('.', '').replace(',', '.')

                match_numero = re.search(r'NÚMERO:\s*(\d{3}\.\d{3}\.\d{3}-\w{2})', text)
                if match_numero:
                    numero_fatura = match_numero.group(1)
                    valores_totais[numero_fatura] = valor_total

        return valores_totais

def executar_codigo_secundario():
    arquivo_rar = "Faturas TIM.rar"
    pasta_destino = os.getcwd()

    if not os.path.exists(arquivo_rar):
        print(f"Arquivo {arquivo_rar} não encontrado.")
    else:
        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)
        caminho_arquivo_txt = extrair_arquivo_rar(arquivo_rar, pasta_destino)
        if caminho_arquivo_txt:
            fatura_xlsx = converter_para_xlsx(caminho_arquivo_txt)
        else:
            print("Arquivo fatura.txt não encontrado na pasta descompactada.")
            exit(1)

    fatura_pdf_path = os.path.join(pasta_destino, "Faturas TIM", "fatura.pdf")

    if not os.path.exists(fatura_pdf_path):
        print("Arquivo fatura.pdf não encontrado. Certifique-se de que ele esteja no mesmo diretório do script.")
        exit(1)

    try:
        df = pd.read_excel(fatura_xlsx, engine='openpyxl')
        num_nf_list = df['NumNF'].tolist()
    except FileNotFoundError:
        print("Arquivo fatura.xlsx não encontrado. Certifique-se de que ele esteja no mesmo diretório do script.")
        exit(1)

    matching_data = []
    try:
        with open(fatura_pdf_path, 'rb') as pdf_file:
            pdf_reader = PdfReader(pdf_file)
            for num_nf in num_nf_list:
                for page_num, page in enumerate(pdf_reader.pages, start=1):
                    pdf_text = page.extract_text()
                    if num_nf in pdf_text:
                        cnpjs = extract_cnpj(pdf_text)
                        if cnpjs:
                            matching_data.append([num_nf, cnpjs[0]]) 
                        break 
    except FileNotFoundError:
        print("Arquivo fatura.pdf não encontrado. Certifique-se de que ele esteja no mesmo diretório do script.")
        exit(1)

    workbook = openpyxl.load_workbook(fatura_xlsx)
    sheet = workbook.active
    faturas = [str(cell.value) for cell in sheet['A'] if cell.value is not None]

    faturas_encontradas = extract_faturas_cnpj_from_pdf(fatura_pdf_path, faturas)

    for i, item in enumerate(matching_data):
        for fatura, cnpj in faturas_encontradas:
            if item[0] == fatura:
                matching_data[i][1] = cnpj

    output_csv_filename = os.path.join(pasta_destino, 'dados_correspondentes.csv')
    try:
        with open(output_csv_filename, 'w', newline='') as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(['NumNF', 'CNPJ'])
            writer.writerows(matching_data)
        print(f"Os dados correspondentes foram salvos em {output_csv_filename}")
    except PermissionError:
        print(f"Não foi possível salvar o arquivo {output_csv_filename}. Verifique as permissões do diretório.")
        exit(1)

    dados_correspondentes = pd.read_csv(output_csv_filename)

    fatura = pd.read_excel(fatura_xlsx, engine='openpyxl')

    fatura_convertida = pd.merge(fatura, dados_correspondentes, on='NumNF', how='left')

    caminho_arquivo_convertido = os.path.join(pasta_destino, "convertidos.xlsx")
    fatura_convertida.to_excel(caminho_arquivo_convertido, index=False)

    print(f"O arquivo convertidos.xlsx foi criado em: {caminho_arquivo_convertido}")

    os.remove(output_csv_filename)
    os.remove(fatura_xlsx)

    pasta_destino_faturas_tim = os.path.join(pasta_destino, 'Faturas TIM')
    for arquivo in os.listdir(pasta_destino_faturas_tim):
        if arquivo != 'convertidos.xlsx' and arquivo != 'fatura.pdf':
            os.remove(os.path.join(pasta_destino_faturas_tim, arquivo))

    print("Arquivos temporários excluídos e pasta de destino limpa.")

    caminho_pdf = os.path.join(pasta_destino, "Faturas TIM", 'fatura.pdf')

    valores_totais = extrair_valores(caminho_pdf)

    if valores_totais:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet.cell(row=1, column=1, value="FNF")
        sheet.cell(row=1, column=2, value="NumNF")

        for i, (numero_fatura, valor_total) in enumerate(valores_totais.items(), start=2):
            valor_total = valor_total.replace(',', '.') 
            sheet.cell(row=i, column=1, value=valor_total)
            sheet.cell(row=i, column=2, value=numero_fatura)

        caminho_xlsx = "valores_totais_com_numeros.xlsx"
        workbook.save(os.path.join(pasta_destino, caminho_xlsx))

        print(f"Os valores totais e números de fatura foram salvos em '{caminho_xlsx}'.")

        convertidos = pd.read_excel(os.path.join(pasta_destino, "convertidos.xlsx"))
        valores_totais = pd.read_excel(os.path.join(pasta_destino, caminho_xlsx))

        final = pd.merge(convertidos, valores_totais[['NumNF', 'FNF']], on='NumNF', how='left')

        final['SNF'] = final['Valor'] - final['FNF']

        caminho_final = os.path.join(pasta_destino, "final.xlsx")
        final.to_excel(caminho_final, index=False)

        print("Os dados foram unificados e salvos em 'final.xlsx'.")

        os.remove(os.path.join(pasta_destino, caminho_xlsx))
        os.remove(caminho_arquivo_convertido)
        shutil.rmtree(os.path.join(pasta_destino, 'Faturas TIM'))

        print("Arquivos temporários e pasta 'Faturas TIM' removidos.")
    else:
        print("Não foi possível encontrar nenhum valor total.")

if __name__ == "__main__":
    executar_codigo_secundario() 
